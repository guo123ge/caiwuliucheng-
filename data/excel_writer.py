from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
SUBTITLE_FONT = Font(name="微软雅黑", size=11, bold=True)
SUMMARY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")


def _auto_width(ws, min_width: int = 8, max_width: int = 40):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    length = 0
                    for ch in line:
                        length += 2 if ord(ch) > 127 else 1
                    max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _apply_header_style(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _apply_data_style(ws, start_row: int, end_row: int, col_count: int,
                       number_cols: Optional[List[int]] = None):
    if number_cols is None:
        number_cols = []
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col in number_cols:
                cell.alignment = RIGHT_ALIGN
                if cell.value is not None:
                    try:
                        cell.number_format = "#,##0.00"
                    except Exception:
                        pass
            else:
                cell.alignment = LEFT_ALIGN


def write_dataframe_to_sheet(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    title: Optional[str] = None,
    number_cols: Optional[List[int]] = None,
    summary_rows: Optional[List[int]] = None,
    warning_rows: Optional[List[int]] = None,
):
    if title:
        ws.merge_cells(start_row=start_row, start_column=1,
                        end_row=start_row, end_column=len(df.columns))
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN
        start_row += 1

    header_row = start_row
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=header_row, column=col_idx, value=str(col_name))
    _apply_header_style(ws, header_row, len(df.columns))

    data_start = header_row + 1
    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        excel_row = data_start + row_idx
        for col_idx, col_name in enumerate(df.columns, 1):
            value = row_data[col_name]
            if pd.isna(value):
                value = ""
            ws.cell(row=excel_row, column=col_idx, value=value)

    data_end = data_start + len(df) - 1
    _apply_data_style(ws, data_start, data_end, len(df.columns), number_cols)

    if summary_rows:
        for sr in summary_rows:
            actual_row = data_start + sr if sr >= 0 else data_end + 1 + sr
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=actual_row, column=col)
                cell.fill = SUMMARY_FILL
                cell.font = SUBTITLE_FONT
                cell.border = THIN_BORDER

    if warning_rows:
        for wr in warning_rows:
            actual_row = data_start + wr if wr >= 0 else data_end + 1 + wr
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=actual_row, column=col)
                cell.fill = WARNING_FILL
                cell.border = THIN_BORDER

    _auto_width(ws)


def create_workbook_with_sheets(
    sheets: List[Dict[str, Any]],
    output_path: str,
):
    wb = Workbook()
    wb.remove(wb.active)

    for i, sheet_spec in enumerate(sheets):
        if i == 0:
            ws = wb.create_sheet(title=sheet_spec.get("name", "Sheet1"), index=0)
        else:
            ws = wb.create_sheet(title=sheet_spec.get("name", f"Sheet{i+1}"))

        df = sheet_spec.get("dataframe")
        if df is not None:
            write_dataframe_to_sheet(
                ws,
                df,
                title=sheet_spec.get("title"),
                number_cols=sheet_spec.get("number_cols"),
                summary_rows=sheet_spec.get("summary_rows"),
                warning_rows=sheet_spec.get("warning_rows"),
            )

    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def save_dataframe(df: pd.DataFrame, output_path: str, title: Optional[str] = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    number_cols = []
    for i, col in enumerate(df.columns, 1):
        if df[col].dtype in ("float64", "int64"):
            number_cols.append(i)

    write_dataframe_to_sheet(ws, df, title=title, number_cols=number_cols)

    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
